from log_utils import logger_print

# first get the titles.
# 解析设备参数表 可能也适用于设备接口信息解析
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pandas
import rich
import numpy
import json

import os

if os.name == "nt":
    from win32com.client import Dispatch

    def repair_excel(excel_path):  # you may need to restart system if this goes wrong.
        xlapp = Dispatch("ket.Application")  # wps
        # xlapp = Dispatch("Excel.Application")
        xlapp.Visible = False
        xlbook = xlapp.Workbooks.Open(os.path.abspath(excel_path))
        xlbook.Save()
        xlbook.Close()

else:
    # in macos
    def repair_excel(excel_path):
        import tempfile

        soffice_bin = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
        with tempfile.TemporaryDirectory() as TD:
            tmpdir = os.path.abspath(TD)
            excel_path_abs = os.path.abspath(excel_path)
            excel_path_base = os.path.basename(excel_path)
            commandline = f"'{soffice_bin}' --headless --convert-to 'xlsx:{excel_path_base.split('.')[0]}' --outdir '{tmpdir}' '{excel_path_abs}'"
            os.system(commandline)
            os.system(f"mv {os.path.join(tmpdir, excel_path_base)} {excel_path_abs}")


def strip_and_convert_empty_string_as_none(e: str):
    stripped = e.strip()
    if stripped == "":
        return None
    else:
        return stripped


def strip_element_if_is_string(mlist):
    return [
        e if not isinstance(e, str) else strip_and_convert_empty_string_as_none(e)
        for e in mlist
    ]


def main_parser(filepath, sheet_name, output_path, type_utils_parser: bool):
    # if os.name == "nt":
    sheet1 = extract_sheet_from_excel(filepath, sheet_name)
    # common parts?
    # order: category; name (unit), example, delete or not
    # you need to scan through all cells to find some cell with specific color.
    # and with some example.
    # COL: A;B,C,D;F,G,H for all data need to export

    # after (partial) serialization, you can do something more interesting with it.
    dims = sheet1.row_dimensions, sheet1.column_dimensions
    uniqs = {}

    def getColumnRangePerRow(start, end):
        flag = True
        for index, row in enumerate(sheet1.rows):
            if flag:
                flag = False
                continue
            yield index, [col.value for col in row[start:end]]

    heads = getColumnRangePerRow(0, 1)
    unwanted_headers = ["其他"]
    # cursor = None
    headMaps = {}
    prevHead = None
    mHeads = []
    for index, [head] in heads:
        logger_print(index, head)
        if head:
            prevHead = head
            mHeads.append(head)
        if prevHead:
            headMaps.update({index: prevHead})

    logger_print(headMaps)
    target_json = {
        h: {}
        for h in mHeads
        if (True if type_utils_parser is False else h not in unwanted_headers)
    }

    if type_utils_parser:
        # breakpoint()

        deviceNameAndPorts = list(getColumnRangePerRow(1, 3))
        deviceNameAndPortsWithoutIndex = [row for _, row in deviceNameAndPorts]
        details = list(getColumnRangePerRow(4, 10))
        indexs = [i for i, _ in details]
        details_without_index = [row for _, row in details]
        # breakpoint()
        currentDevName = None
        currentDevData = None
        prevDevName = None
        createDevDataTemplate = lambda: {"ports": {}, "rules": [], "requirements": []}
        for i, row_i in enumerate(indexs):
            head = headMaps[row_i]
            if head in unwanted_headers:
                logger_print(
                    f'skipping row #{row_i} because of "{head}" is in unwanted headers ({unwanted_headers}).'
                )
                continue
            devNameOrPortName, portInfo = strip_element_if_is_string(
                deviceNameAndPortsWithoutIndex[i]
            )
            细分类型, 基本类型, 能流方向, 必有工况, 对应规则, 附加要求 = strip_element_if_is_string(
                details_without_index[i]
            )

            if currentDevName is not None:
                if devNameOrPortName is None:
                    currentDevName = None
                else:
                    if currentDevData is None:
                        currentDevData = createDevDataTemplate()
                    portName = devNameOrPortName
                    currentDevData["ports"][portName] = dict(
                        info=portInfo, 细分类型=细分类型, 基本类型=基本类型, 能流方向=能流方向, 必有工况=必有工况
                    )
                    if 对应规则 is not None:
                        currentDevData["rules"].append(对应规则)
                    if 附加要求 is not None:
                        currentDevData["requirements"].append(附加要求)
            else:
                if currentDevData is not None:
                    prevHead2 = headMaps[row_i - 1]
                    target_json[prevHead2][prevDevName] = currentDevData
                    currentDevData = None
                if portInfo is None:
                    if devNameOrPortName is not None:
                        currentDevName = devNameOrPortName
                        prevDevName = currentDevName

    else:
        BCD = getColumnRangePerRow(1, 4)
        FGH = getColumnRangePerRow(5, 8)

        def checkEmpty(val):
            if type(val) == str:
                if val.strip() == "":
                    return None
            return val

        def processBCD(_BCD):
            device_name = None
            for index, [b, c, d] in _BCD:
                head = headMaps[index]
                b, c, d = checkEmpty(b), checkEmpty(c), checkEmpty(d)
                # logger_print(b,c,d) # value can be None or ""
                if all([elem is None for elem in [b, c, d]]):
                    logger_print("LINE BREAK")
                    device_name = None
                else:
                    if device_name is None:
                        device_name = b
                        target_json[head].update({device_name: []})
                    else:
                        target_json[head][device_name].append((b, c, d))
                    logger_print("DEVICE NAME?", device_name)

        processBCD(BCD)
        processBCD(FGH)

    logger_print(target_json)

    with open(output_path, "w+") as f:
        f.write(json.dumps(target_json, indent=4, ensure_ascii=False))
    logger_print("WRITE TO:", output_path)


def extract_sheet_from_excel(filepath, sheet_name):
    repair_excel(filepath)

    excel_file = openpyxl.load_workbook(filepath)
    # excel_file = openpyxl.load_workbook(filepath, read_only=True)
    logger_print("SHEET NAMES:")
    logger_print(excel_file.sheetnames)  # ['Sheet1']
    # from openpyxl.cell.cell import Cell, MergedCell

    sheet1 = excel_file[sheet_name]

    if not isinstance(sheet1, Worksheet):
        raise Exception(
            f"sheet {sheet_name} at file '{filepath}' (type: {type(sheet1)}) is not Worksheet"
        )
    return sheet1


def csv_parser(filename, output_path):
    df = pandas.read_csv(filename, header=None)
    dataClasses = [None, None]
    result = {}
    lastEmpty = True
    for index, row in df.iterrows():
        # logger_print(row)
        # logger_print(list(row))
        list_row = list(row)
        first, second = list_row[:2]
        # logger_print(dir(row))
        if first is numpy.nan and second is numpy.nan:
            lastEmpty = True
            continue
        # list_row_types = [(e, type(e)) for e in list_row]
        # logger_print(list_row_types)
        # numpy.nan is a float, not an int, so we can't use it as a number
        if type(first) == str:
            first = first.strip()
            if len(first) > 0:
                dataClasses[0] = first
        if type(second) == str:
            second = second.strip()
            if len(second) > 0:
                if lastEmpty:
                    dataClasses[1] = second
                    lastEmpty = False
                else:
                    # now we begin to insert data.
                    if dataClasses[0] and dataClasses[1]:
                        if result.get(dataClasses[0], None) is None:
                            result[dataClasses[0]] = {}
                        if result[dataClasses[0]].get(dataClasses[1], None) is None:
                            result[dataClasses[0]][dataClasses[1]] = []
                        result[dataClasses[0]][dataClasses[1]].append(second)
    logger_print(result)
    with open(output_path, "w+") as f:
        f.write(json.dumps(result, indent=4, ensure_ascii=False))

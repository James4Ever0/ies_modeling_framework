# first get the titles.
import openpyxl

filepath = "设备信息库各参数.xlsx"

excel_file = openpyxl.load_workbook(filepath)

sheet1 = excel_file["设备参数"]



import openpyxl

filepath = "device_parameters_v3.3.xlsx"

excel_file = openpyxl.load_workbook(filename=filepath)
# print(excel_file.sheetnames) # ['Sheet1']
from openpyxl.worksheet.worksheet import Worksheet
# from openpyxl.cell.cell import Cell, MergedCell

sheet1 = excel_file["Sheet1"]
if type(sheet1) == Worksheet:
    # order: category; name (unit), example, delete or not
    # you need to scan through all cells to find some cell with specific color.
    # and with some example.
    # COL: A;B,C,D;F,G,H for all data need to export

    # after (partial) serialization, you can do something more interesting with it.
    dims = sheet1.row_dimensions, sheet1.column_dimensions
    # print(dims)
    # breakpoint()
    # print(sheet1)
    # print(type(sheet1))
    # breakpoint()
    # print(dir(sheet1))
    # breakpoint()
    uniqs = {}

    def getColumnRangePerRow(start, end):
        flag = True
        for index, row in enumerate(sheet1.rows):
            if flag:
                flag = False
                continue
            yield index, [col.value for col in row[start:end]]

    heads = getColumnRangePerRow(0, 1)
    # cursor = None
    headMaps = {}
    prevHead = None
    mHeads = []
    for index, [head] in heads:
        print(index, head)
        if head:
            # print(type(head))
            prevHead = head
            mHeads.append(head)
        if prevHead:
            headMaps.update({index: prevHead})
    import rich

    rich.print(headMaps)

    BCD = getColumnRangePerRow(1, 4)
    FGH = getColumnRangePerRow(5, 8)

    def checkEmpty(val):
        if type(val) == str:
            if val.strip() == "":
                return None
        return val

    target_json = {h: {} for h in mHeads}

    def processBCD(_BCD):
        device_name = None
        for index, [b, c, d] in _BCD:
            head = headMaps[index]
            b, c, d = checkEmpty(b), checkEmpty(c), checkEmpty(d)
            # print(b,c,d) # value can be None or ""
            if all([elem is None for elem in [b, c, d]]):
                print("LINE BREAK")
                device_name = None
            else:
                if device_name is None:
                    device_name = b
                    target_json[head].update({device_name: []})
                else:
                    target_json[head][device_name].append((b, c, d))
                print("DEVICE NAME?", device_name)

    processBCD(BCD)
    processBCD(FGH)

    rich.print(target_json)
    import json

    output_path = "device_params_intermediate.json"
    with open(output_path, "w+") as f:
        f.write(json.dumps(target_json, indent=4, ensure_ascii=False))
    print("WRITE TO:", output_path)

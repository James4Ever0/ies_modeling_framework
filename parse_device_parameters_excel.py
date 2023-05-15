import openpyxl

filepath = "device_parameters_v3.3.xlsx"

excel_file = openpyxl.load_workbook(filename=filepath)
# print(excel_file.sheetnames) # ['Sheet1']
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell 

sheet1 = excel_file["Sheet1"]
if type(sheet1) == Worksheet:
    # order: category; name (unit), example, delete or not
    # you need to scan through all cells to find some cell with specific color.
    # and with some example.
    # COL: A;B,C,D;F,G,H for all data need to export
    
    # after (partial) serialization, you can do something more interesting with it.
    dims = sheet1.row_dimensions, sheet1.column_dimensions
    # print(dims)
    # breakpoint()
    # print(sheet1)
    # print(type(sheet1))
    # breakpoint()
    # print(dir(sheet1))
    # breakpoint()
    uniqs = {}
    for row in sheet1.rows:
        # a tuple containing every cell in the row
        for cell1 in row:
            # cell1 = sheet1.cell(row=1, column=1)  # cell or merged cell.
            # need to determine its type.
            # now we don't care about the color anymore.
            # we just want the value.
            if type(cell1) == Cell:
                # print(type(cell1))
                # print(dir(cell1))
                # breakpoint()
                cell1_fill = cell1.fill
                cell1_value = cell1.value
                cell1_column_letter = cell1.column_letter
                cell1_font = cell1.font
                # breakpoint()
                print()
                print("FONT:",cell1_font) # font color here.
                print()
                print("FILL:", cell1_fill) # includes fg and bg
                print()
                print("VALUE:", cell1_value)
                if cell1_value in [
                    # '增加', # fg: FF92D050
                                #    '水水换热器',
                                #    '非必填', (no color?)
                                #    '配电传输' # fg: FF92D050
                                   ]:
                    # breakpoint()
                    ...
                
                font_color = cell1_font.color
                fgColor = cell1_fill.fgColor
                bgColor = cell1_fill.bgColor
                
                if font_color:
                    uniqs.update({font_color.rgb: cell1_value})
                if fgColor:
                    uniqs.update({fgColor.rgb: cell1_value})
                if bgColor:
                    uniqs.update({bgColor.rgb: cell1_value})
                # breakpoint()
                # use .rgb to access the color string
                # '00000000'
            elif type(cell1) == MergedCell:
                ...
            else:
                print("Unknown cell type: %s" % type(cell1))
    print("*" * 50)
    for key, value in uniqs.items():
        print(key, value)
# 00000000 None
# FF92D050 计算单位功率成本
# FFFFFF00 气水换热器
# FFFF0000 None